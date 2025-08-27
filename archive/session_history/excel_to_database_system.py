"""
Excel to Database System - Handle unstructured Excel files intelligently
Extract data, store in database, make it searchable with AI assistance
"""

import os
import sys
import json
import asyncio
import sqlite3
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple

import openpyxl
from openpyxl import load_workbook
import httpx

# Database and file paths - Dynamic path detection
import os
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATABASE_PATH = os.path.join(PROJECT_ROOT, "excel_data.db")
EXCEL_FILES_PATH = os.path.join(PROJECT_ROOT, "data", "input")

# LM Studio configuration for AI analysis
LM_STUDIO_BASE_URL = "http://192.168.101.70:1234/v1"
LM_STUDIO_MODEL = "openai/gpt-oss-20b"


class ExcelDatabaseSystem:
    """Comprehensive system for handling unstructured Excel files."""
    
    def __init__(self, db_path: str = DATABASE_PATH):
        self.db_path = db_path
        self.connection = None
        self.setup_database()
    
    def setup_database(self):
        """Create database schema for Excel data storage."""
        try:
            self.connection = sqlite3.connect(self.db_path)
            self.connection.execute("PRAGMA foreign_keys = ON")
            
            # Main files table
            self.connection.execute("""
                CREATE TABLE IF NOT EXISTS excel_files (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    filename TEXT NOT NULL,
                    file_path TEXT NOT NULL,
                    file_size INTEGER,
                    sheets_count INTEGER,
                    processed_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    file_hash TEXT,
                    business_type TEXT,
                    ai_summary TEXT
                )
            """)
            
            # Sheets table
            self.connection.execute("""
                CREATE TABLE IF NOT EXISTS excel_sheets (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    file_id INTEGER,
                    sheet_name TEXT NOT NULL,
                    sheet_index INTEGER,
                    rows_count INTEGER,
                    columns_count INTEGER,
                    data_type TEXT,  -- financial, operational, master_data, etc.
                    ai_interpretation TEXT,
                    FOREIGN KEY (file_id) REFERENCES excel_files (id)
                )
            """)
            
            # Raw data table - stores actual cell values
            self.connection.execute("""
                CREATE TABLE IF NOT EXISTS excel_data (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    sheet_id INTEGER,
                    row_number INTEGER,
                    column_number INTEGER,
                    column_letter TEXT,
                    cell_address TEXT,
                    raw_value TEXT,
                    data_type TEXT,  -- number, text, date, formula, empty
                    numeric_value REAL,
                    text_value TEXT,
                    date_value DATE,
                    is_header BOOLEAN DEFAULT 0,
                    context_info TEXT,
                    FOREIGN KEY (sheet_id) REFERENCES excel_sheets (id)
                )
            """)
            
            # Structured data table - AI-interpreted financial/business data
            self.connection.execute("""
                CREATE TABLE IF NOT EXISTS financial_data (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    sheet_id INTEGER,
                    entity_name TEXT,
                    metric_name TEXT,  -- revenue, profit, assets, etc.
                    metric_category TEXT,  -- income_statement, balance_sheet, cash_flow
                    metric_value REAL,
                    currency TEXT DEFAULT 'RM',
                    period_date DATE,
                    period_type TEXT,  -- monthly, quarterly, yearly
                    confidence_score REAL,  -- AI confidence in extraction
                    extraction_method TEXT,  -- manual, ai_pattern, ai_nlp
                    notes TEXT,
                    FOREIGN KEY (sheet_id) REFERENCES excel_sheets (id)
                )
            """)
            
            # Search index table for fast text search
            self.connection.execute("""
                CREATE TABLE IF NOT EXISTS search_index (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    file_id INTEGER,
                    sheet_id INTEGER,
                    searchable_text TEXT,
                    content_type TEXT,  -- header, data, entity, metric
                    relevance_score REAL DEFAULT 1.0,
                    FOREIGN KEY (file_id) REFERENCES excel_files (id),
                    FOREIGN KEY (sheet_id) REFERENCES excel_sheets (id)
                )
            """)
            
            # Create indexes for better query performance
            self.connection.execute("CREATE INDEX IF NOT EXISTS idx_excel_data_sheet ON excel_data(sheet_id)")
            self.connection.execute("CREATE INDEX IF NOT EXISTS idx_financial_data_entity ON financial_data(entity_name)")
            self.connection.execute("CREATE INDEX IF NOT EXISTS idx_search_text ON search_index(searchable_text)")
            
            self.connection.commit()
            print("✅ Database schema created successfully")
            
        except Exception as e:
            print(f"❌ Database setup failed: {e}")
            raise
    
    async def call_ai_analysis(self, prompt: str, context: str = "") -> str:
        """Call LM Studio for AI analysis of Excel data."""
        try:
            messages = [
                {
                    "role": "system",
                    "content": """You are an AI expert in analyzing unstructured Excel files for business and financial data. 
                    
Provide structured analysis in JSON format when requested, or clear explanations for data interpretation."""
                },
                {
                    "role": "user", 
                    "content": f"{context}\n\n{prompt}"
                }
            ]
            
            async with httpx.AsyncClient(timeout=60.0) as client:
                response = await client.post(
                    f"{LM_STUDIO_BASE_URL}/chat/completions",
                    json={
                        "model": LM_STUDIO_MODEL,
                        "messages": messages,
                        "temperature": 0.2,
                        "max_tokens": 2000,
                    },
                    headers={"Content-Type": "application/json"}
                )
                
                if response.status_code == 200:
                    data = response.json()
                    return data["choices"][0]["message"]["content"]
                else:
                    return f"AI analysis failed: HTTP {response.status_code}"
                    
        except Exception as e:
            return f"AI analysis error: {str(e)}"
    
    def extract_excel_data(self, file_path: str) -> Dict[str, Any]:
        """Extract all data from Excel file with intelligent analysis."""
        try:
            print(f"📖 Processing Excel file: {file_path}")
            
            workbook = load_workbook(file_path, read_only=True)
            file_info = {
                "filename": Path(file_path).name,
                "file_path": file_path,
                "file_size": os.path.getsize(file_path),
                "sheets_count": len(workbook.sheetnames),
                "sheets_data": {}
            }
            
            for sheet_index, sheet_name in enumerate(workbook.sheetnames):
                print(f"   📋 Processing sheet: {sheet_name}")
                worksheet = workbook[sheet_name]
                
                # Extract all cell data
                sheet_data = {
                    "sheet_name": sheet_name,
                    "sheet_index": sheet_index,
                    "rows_count": worksheet.max_row,
                    "columns_count": worksheet.max_column,
                    "cells": [],
                    "headers": [],
                    "numeric_data": [],
                    "text_data": [],
                    "patterns": {
                        "financial_terms": [],
                        "entities": [],
                        "dates": [],
                        "currencies": []
                    }
                }
                
                # Read cells more efficiently - focus on data-rich areas
                max_rows_to_read = min(worksheet.max_row, 50)  # Reduced from 200 to 50 rows
                max_cols_to_read = min(worksheet.max_column, 20)  # Reduced from 50 to 20 columns

                # First, try to find the data range by scanning for non-empty cells
                data_rows = []
                for row_num in range(1, min(worksheet.max_row + 1, max_rows_to_read)):
                    row_has_data = False
                    for col_num in range(1, min(worksheet.max_column + 1, max_cols_to_read)):
                        cell = worksheet.cell(row=row_num, column=col_num)
                        if cell.value is not None and str(cell.value).strip():
                            row_has_data = True
                            break
                    if row_has_data:
                        data_rows.append(row_num)

                # Read only rows that have data
                for row_num in data_rows[:30]:  # Limit to 30 data-rich rows
                    for col_num in range(1, min(worksheet.max_column + 1, max_cols_to_read)):
                        cell = worksheet.cell(row=row_num, column=col_num)
                        cell_value = cell.value

                        if cell_value is not None and str(cell_value).strip():
                            column_letter = openpyxl.utils.get_column_letter(col_num)
                            cell_address = f"{column_letter}{row_num}"
                            
                            # Determine data type and extract value
                            cell_info = {
                                "row": row_num,
                                "column": col_num,
                                "column_letter": column_letter,
                                "address": cell_address,
                                "raw_value": str(cell_value),
                                "data_type": self._determine_data_type(cell_value),
                                "numeric_value": None,
                                "text_value": None,
                                "date_value": None,
                                "is_header": row_num <= 5  # Assume first 5 rows might be headers
                            }
                            
                            # Extract typed values
                            if isinstance(cell_value, (int, float)):
                                cell_info["numeric_value"] = float(cell_value)
                                sheet_data["numeric_data"].append(cell_value)
                            elif isinstance(cell_value, str) and cell_value.strip():
                                cell_info["text_value"] = cell_value.strip()
                                sheet_data["text_data"].append(cell_value.strip())
                                
                                # Pattern recognition
                                self._extract_patterns(cell_value.strip(), sheet_data["patterns"])
                            
                            sheet_data["cells"].append(cell_info)
                            
                            # Collect potential headers
                            if row_num <= 10 and isinstance(cell_value, str) and len(cell_value.strip()) > 2:
                                sheet_data["headers"].append(cell_value.strip())
                
                file_info["sheets_data"][sheet_name] = sheet_data
            
            workbook.close()
            return file_info
            
        except Exception as e:
            return {"error": f"Failed to extract Excel data: {str(e)}"}
    
    def _determine_data_type(self, value) -> str:
        """Determine the data type of a cell value."""
        if isinstance(value, (int, float)):
            return "number"
        elif isinstance(value, str):
            if value.strip() == "":
                return "empty"
            elif value.startswith("="):
                return "formula"
            else:
                return "text"
        elif hasattr(value, 'date'):  # datetime object
            return "date"
        else:
            return "unknown"
    
    def _extract_patterns(self, text: str, patterns: Dict[str, List[str]]):
        """Extract business patterns from text."""
        text_lower = text.lower()
        
        # Financial terms
        financial_keywords = [
            'revenue', 'profit', 'loss', 'assets', 'liabilities', 'equity', 
            'cash', 'expenses', 'income', 'ebitda', 'gross profit', 'net profit',
            'balance', 'total', 'turnover', 'cost', 'margin'
        ]
        
        for keyword in financial_keywords:
            if keyword in text_lower and text not in patterns["financial_terms"]:
                patterns["financial_terms"].append(text[:100])
                break
        
        # Malaysian business entities
        entity_keywords = ['sdn bhd', 'berhad', 'bhd', 'corporation', 'holdings', 'group']
        for keyword in entity_keywords:
            if keyword in text_lower and text not in patterns["entities"]:
                patterns["entities"].append(text[:100])
                break
        
        # Currency mentions
        currency_keywords = ['rm ', 'ringgit', 'million', 'thousand', 'usd', 'sgd']
        for keyword in currency_keywords:
            if keyword in text_lower and text not in patterns["currencies"]:
                patterns["currencies"].append(text[:100])
                break
    
    async def store_in_database(self, file_data: Dict[str, Any]) -> int:
        """Store extracted Excel data in database with AI analysis."""
        try:
            cursor = self.connection.cursor()
            
            # Get AI analysis of the file (optional - don't fail if LM Studio is unavailable)
            ai_summary = "Excel file processed successfully"
            try:
                ai_context = self._build_ai_context(file_data)
                ai_summary = await self.call_ai_analysis(
                    "Analyze this Excel file and provide a brief summary of its business purpose and data types.",
                    ai_context
                )
            except Exception as e:
                print(f"⚠️  AI analysis failed (continuing without it): {e}")
                ai_summary = f"Excel file processed successfully (AI analysis unavailable: {str(e)})"
            
            # Insert file record
            cursor.execute("""
                INSERT INTO excel_files (filename, file_path, file_size, sheets_count, ai_summary)
                VALUES (?, ?, ?, ?, ?)
            """, (
                file_data["filename"],
                file_data["file_path"], 
                file_data["file_size"],
                file_data["sheets_count"],
                ai_summary[:1000]  # Limit summary length
            ))
            
            file_id = cursor.lastrowid
            print(f"📁 Stored file record: ID {file_id}")
            
            # Process each sheet
            for sheet_name, sheet_data in file_data["sheets_data"].items():
                # Get AI interpretation of sheet
                sheet_context = f"""Sheet: {sheet_name}
Headers found: {', '.join(sheet_data['headers'][:10])}
Financial terms: {', '.join(sheet_data['patterns']['financial_terms'][:5])}
Entities: {', '.join(sheet_data['patterns']['entities'][:3])}
Numeric data count: {len(sheet_data['numeric_data'])}"""

                sheet_interpretation = await self.call_ai_analysis(
                    f"What type of business data is in this sheet? (financial_statement, operational_data, master_data, etc.)",
                    sheet_context
                )
                
                # Insert sheet record
                cursor.execute("""
                    INSERT INTO excel_sheets (file_id, sheet_name, sheet_index, rows_count, columns_count, data_type, ai_interpretation)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                """, (
                    file_id,
                    sheet_name,
                    sheet_data["sheet_index"],
                    sheet_data["rows_count"],
                    sheet_data["columns_count"],
                    sheet_interpretation[:50],  # First 50 chars as data type
                    sheet_interpretation[:500]
                ))
                
                sheet_id = cursor.lastrowid
                print(f"   📋 Stored sheet record: {sheet_name} (ID {sheet_id})")
                
                # Insert cell data in batches
                cell_batch = []
                for cell in sheet_data["cells"]:
                    cell_batch.append((
                        sheet_id,
                        cell["row"],
                        cell["column"], 
                        cell["column_letter"],
                        cell["address"],
                        cell["raw_value"][:500],  # Limit text length
                        cell["data_type"],
                        cell["numeric_value"],
                        cell["text_value"][:200] if cell["text_value"] else None,
                        cell["date_value"],
                        cell["is_header"]
                    ))
                    
                    # Insert in batches of 1000
                    if len(cell_batch) >= 1000:
                        cursor.executemany("""
                            INSERT INTO excel_data (sheet_id, row_number, column_number, column_letter, 
                                                   cell_address, raw_value, data_type, numeric_value, 
                                                   text_value, date_value, is_header)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """, cell_batch)
                        cell_batch = []
                
                # Insert remaining cells
                if cell_batch:
                    cursor.executemany("""
                        INSERT INTO excel_data (sheet_id, row_number, column_number, column_letter, 
                                               cell_address, raw_value, data_type, numeric_value, 
                                               text_value, date_value, is_header)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, cell_batch)
                
                # Build search index
                search_entries = []
                
                # Add headers to search
                for header in sheet_data["headers"]:
                    search_entries.append((file_id, sheet_id, header, "header", 1.0))
                
                # Add patterns to search
                for pattern_type, patterns in sheet_data["patterns"].items():
                    for pattern in patterns:
                        search_entries.append((file_id, sheet_id, pattern, pattern_type, 0.8))
                
                if search_entries:
                    cursor.executemany("""
                        INSERT INTO search_index (file_id, sheet_id, searchable_text, content_type, relevance_score)
                        VALUES (?, ?, ?, ?, ?)
                    """, search_entries)
                
                print(f"     💾 Stored {len(sheet_data['cells'])} cells and {len(search_entries)} search entries")
            
            self.connection.commit()
            print(f"✅ Successfully stored Excel file in database (File ID: {file_id})")
            return file_id
            
        except Exception as e:
            self.connection.rollback()
            print(f"❌ Database storage failed: {e}")
            raise
    
    def _build_ai_context(self, file_data: Dict[str, Any]) -> str:
        """Build context string for AI analysis."""
        context = f"""Excel File Analysis:
Filename: {file_data['filename']}
Sheets: {file_data['sheets_count']}
Sheet Names: {', '.join(file_data['sheets_data'].keys())}

Sample Data:"""
        
        for sheet_name, sheet_data in list(file_data["sheets_data"].items())[:2]:  # First 2 sheets
            context += f"\n\nSheet '{sheet_name}':"
            context += f"\nHeaders: {', '.join(sheet_data['headers'][:5])}"
            context += f"\nFinancial terms: {', '.join(sheet_data['patterns']['financial_terms'][:3])}"
            context += f"\nEntities: {', '.join(sheet_data['patterns']['entities'][:2])}"
        
        return context[:2000]  # Limit context size
    
    def search_data(self, query: str, limit: int = 20) -> List[Dict[str, Any]]:
        """Search stored Excel data."""
        try:
            cursor = self.connection.cursor()
            
            # Search in multiple tables
            results = []
            
            # Search in search index
            cursor.execute("""
                SELECT DISTINCT f.filename, s.sheet_name, si.searchable_text, si.content_type, si.relevance_score
                FROM search_index si
                JOIN excel_files f ON si.file_id = f.id  
                JOIN excel_sheets s ON si.sheet_id = s.id
                WHERE si.searchable_text LIKE ?
                ORDER BY si.relevance_score DESC, f.processed_at DESC
                LIMIT ?
            """, (f"%{query}%", limit))
            
            for row in cursor.fetchall():
                results.append({
                    "type": "search_match",
                    "filename": row[0],
                    "sheet_name": row[1], 
                    "matched_text": row[2],
                    "content_type": row[3],
                    "relevance": row[4]
                })
            
            # Search in raw data
            cursor.execute("""
                SELECT DISTINCT f.filename, s.sheet_name, ed.cell_address, ed.raw_value, ed.data_type
                FROM excel_data ed
                JOIN excel_sheets s ON ed.sheet_id = s.id
                JOIN excel_files f ON s.file_id = f.id
                WHERE ed.text_value LIKE ? OR ed.raw_value LIKE ?
                ORDER BY f.processed_at DESC
                LIMIT ?
            """, (f"%{query}%", f"%{query}%", limit))
            
            for row in cursor.fetchall():
                results.append({
                    "type": "cell_match", 
                    "filename": row[0],
                    "sheet_name": row[1],
                    "cell_address": row[2],
                    "cell_value": row[3],
                    "data_type": row[4]
                })
            
            return results[:limit]
            
        except Exception as e:
            print(f"❌ Search failed: {e}")
            return []
    
    async def query_with_ai(self, natural_language_query: str) -> str:
        """Answer questions about stored data using AI."""
        try:
            # Get relevant data based on query
            search_results = self.search_data(natural_language_query, limit=10)
            
            # Build context from search results
            context = "Relevant data from Excel files:\n"
            for result in search_results:
                if result["type"] == "search_match":
                    context += f"- {result['filename']} -> {result['sheet_name']}: {result['matched_text']}\n"
                else:
                    context += f"- {result['filename']} -> {result['sheet_name']} [{result['cell_address']}]: {result['cell_value']}\n"
            
            # Get summary statistics
            cursor = self.connection.cursor()
            cursor.execute("SELECT COUNT(*) FROM excel_files")
            files_count = cursor.fetchone()[0]
            
            cursor.execute("SELECT COUNT(*) FROM financial_data")
            metrics_count = cursor.fetchone()[0]
            
            context += f"\nDatabase Summary: {files_count} Excel files processed, {metrics_count} financial metrics extracted."
            
            # Ask AI to answer the question
            response = await self.call_ai_analysis(
                f"Based on the Excel data provided, please answer this question: {natural_language_query}",
                context
            )
            
            return response
            
        except Exception as e:
            return f"Query failed: {str(e)}"
    
    def get_database_stats(self) -> Dict[str, Any]:
        """Get statistics about stored data."""
        try:
            cursor = self.connection.cursor()
            
            stats = {}
            
            # Files stats
            cursor.execute("SELECT COUNT(*), SUM(file_size), SUM(sheets_count) FROM excel_files")
            files_stats = cursor.fetchone()
            stats["files"] = {
                "count": files_stats[0],
                "total_size_bytes": files_stats[1] or 0,
                "total_sheets": files_stats[2] or 0
            }
            
            # Data stats
            cursor.execute("SELECT COUNT(*) FROM excel_data")
            stats["total_cells"] = cursor.fetchone()[0]
            
            cursor.execute("SELECT COUNT(*) FROM excel_data WHERE data_type = 'number'")
            stats["numeric_cells"] = cursor.fetchone()[0]
            
            cursor.execute("SELECT COUNT(*) FROM search_index")
            stats["searchable_items"] = cursor.fetchone()[0]
            
            # Recent files
            cursor.execute("""
                SELECT filename, sheets_count, processed_at 
                FROM excel_files 
                ORDER BY processed_at DESC 
                LIMIT 5
            """)
            stats["recent_files"] = [
                {"filename": row[0], "sheets": row[1], "processed": row[2]}
                for row in cursor.fetchall()
            ]
            
            return stats
            
        except Exception as e:
            return {"error": f"Stats query failed: {e}"}
    
    def close(self):
        """Close database connection."""
        if self.connection:
            self.connection.close()


async def main():
    """Main function to demonstrate the Excel to Database system."""
    print("🏢 EXCEL TO DATABASE SYSTEM")
    print("=" * 50)
    print("📊 Intelligent Excel processing with database storage and AI search")
    print()
    
    # Initialize system
    system = ExcelDatabaseSystem()
    
    try:
        # Process Excel files
        excel_files_path = Path(EXCEL_FILES_PATH)
        if excel_files_path.exists():
            excel_files = list(excel_files_path.glob("*.xlsx"))
            
            if excel_files:
                print(f"📁 Found {len(excel_files)} Excel files to process:")
                for file_path in excel_files:
                    print(f"   📄 {file_path.name}")
                
                # Process each file
                for file_path in excel_files:
                    print(f"\n🔄 Processing: {file_path.name}")
                    
                    # Extract data
                    file_data = system.extract_excel_data(str(file_path))
                    
                    if "error" not in file_data:
                        # Store in database
                        file_id = await system.store_in_database(file_data)
                        print(f"✅ Successfully processed and stored file (ID: {file_id})")
                    else:
                        print(f"❌ Error processing file: {file_data['error']}")
            else:
                print("⚠️  No Excel files found in input directory")
        else:
            print(f"⚠️  Input directory not found: {excel_files_path}")
        
        # Show database statistics
        print("\n📊 DATABASE STATISTICS:")
        print("-" * 30)
        stats = system.get_database_stats()
        if "error" not in stats:
            print(f"📁 Files processed: {stats['files']['count']}")
            print(f"📋 Total sheets: {stats['files']['total_sheets']}")
            print(f"💾 Total cells stored: {stats['total_cells']:,}")
            print(f"🔢 Numeric cells: {stats['numeric_cells']:,}")
            print(f"🔍 Searchable items: {stats['searchable_items']:,}")
            
            if stats['recent_files']:
                print(f"\n📄 Recent files:")
                for file_info in stats['recent_files']:
                    print(f"   - {file_info['filename']} ({file_info['sheets']} sheets)")
        
        # Demonstrate search functionality
        print("\n🔍 SEARCH DEMONSTRATION:")
        print("-" * 30)
        
        search_queries = ["revenue", "ABC Construction", "profit", "assets"]
        
        for query in search_queries:
            print(f"\nSearching for: '{query}'")
            results = system.search_data(query, limit=5)
            
            if results:
                for i, result in enumerate(results[:3], 1):
                    if result["type"] == "search_match":
                        print(f"   {i}. {result['filename']} -> {result['sheet_name']}: {result['matched_text']}")
                    else:
                        print(f"   {i}. {result['filename']} -> {result['cell_address']}: {result['cell_value']}")
            else:
                print(f"   No results found for '{query}'")
        
        # AI Query demonstration
        print("\n🤖 AI-POWERED QUERY DEMONSTRATION:")
        print("-" * 40)
        
        ai_queries = [
            "What is the total revenue across all files?",
            "Which companies are mentioned in the data?",
            "What financial metrics are available?"
        ]
        
        for query in ai_queries:
            print(f"\nQ: {query}")
            response = await system.query_with_ai(query)
            print(f"A: {response[:200]}{'...' if len(response) > 200 else ''}")
        
    except Exception as e:
        print(f"❌ System error: {e}")
    
    finally:
        system.close()
        print("\n✅ Excel to Database System demonstration complete!")


if __name__ == "__main__":
    asyncio.run(main())