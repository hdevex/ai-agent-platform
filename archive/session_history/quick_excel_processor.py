#!/usr/bin/env python3
"""
Quick Excel Processor - Fast processing without AI analysis
"""

import sqlite3
import openpyxl
from pathlib import Path
import time

DATABASE_PATH = "/mnt/c/Users/nvntr/Documents/ai_agent_platform/excel_data.db"
EXCEL_FILE_PATH = "/mnt/c/Users/nvntr/Documents/ai_agent_platform/data/input/multi_sheet.xlsx"

def quick_process():
    """Quick processing without AI calls."""
    print("⚡ QUICK EXCEL PROCESSING")
    print("=" * 30)
    
    if not Path(EXCEL_FILE_PATH).exists():
        print(f"❌ File not found: {EXCEL_FILE_PATH}")
        return
    
    print(f"📁 Processing: {Path(EXCEL_FILE_PATH).name}")
    
    # Connect to database
    conn = sqlite3.connect(DATABASE_PATH)
    cursor = conn.cursor()
    
    try:
        # Load workbook
        workbook = openpyxl.load_workbook(EXCEL_FILE_PATH, read_only=True)
        
        # Insert file record
        cursor.execute("""
            INSERT INTO excel_files (filename, file_path, file_size, sheets_count, ai_summary)
            VALUES (?, ?, ?, ?, ?)
        """, (
            Path(EXCEL_FILE_PATH).name,
            EXCEL_FILE_PATH,
            Path(EXCEL_FILE_PATH).stat().st_size,
            len(workbook.sheetnames),
            "Quick processed - Malaysian investment holdings Excel file"
        ))
        
        file_id = cursor.lastrowid
        print(f"📁 File ID: {file_id}")
        
        # Process sheets
        for sheet_index, sheet_name in enumerate(workbook.sheetnames):
            print(f"   📋 Processing sheet: {sheet_name}")
            worksheet = workbook[sheet_name]
            
            # Insert sheet record
            cursor.execute("""
                INSERT INTO excel_sheets (file_id, sheet_name, sheet_index, rows_count, columns_count, data_type, ai_interpretation)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (
                file_id, sheet_name, sheet_index, worksheet.max_row, worksheet.max_column,
                "financial_data", f"Excel sheet containing financial data from {sheet_name}"
            ))
            
            sheet_id = cursor.lastrowid
            
            # Process sample data (first 50 rows, 20 columns max)
            cell_batch = []
            search_entries = []
            
            max_rows = min(worksheet.max_row, 50)
            max_cols = min(worksheet.max_column, 20)
            
            for row_num in range(1, max_rows + 1):
                for col_num in range(1, max_cols + 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    if cell.value is not None:
                        column_letter = openpyxl.utils.get_column_letter(col_num)
                        cell_address = f"{column_letter}{row_num}"
                        
                        # Determine data type
                        if isinstance(cell.value, (int, float)):
                            data_type = "number"
                            numeric_value = float(cell.value)
                            text_value = None
                        else:
                            data_type = "text"
                            numeric_value = None
                            text_value = str(cell.value)[:200]
                            
                            # Add to search if it's meaningful text
                            if len(str(cell.value).strip()) > 2:
                                search_entries.append((file_id, sheet_id, str(cell.value)[:100], "data", 1.0))
                        
                        cell_batch.append((
                            sheet_id, row_num, col_num, column_letter, cell_address,
                            str(cell.value)[:500], data_type, numeric_value, text_value,
                            None, row_num <= 3  # First 3 rows as potential headers
                        ))
            
            # Insert cells in batch
            if cell_batch:
                cursor.executemany("""
                    INSERT INTO excel_data (sheet_id, row_number, column_number, column_letter, 
                                           cell_address, raw_value, data_type, numeric_value, 
                                           text_value, date_value, is_header)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, cell_batch)
                
                print(f"      💾 Stored {len(cell_batch)} cells")
            
            # Insert search entries
            if search_entries:
                cursor.executemany("""
                    INSERT INTO search_index (file_id, sheet_id, searchable_text, content_type, relevance_score)
                    VALUES (?, ?, ?, ?, ?)
                """, search_entries[:500])  # Limit search entries
                
                print(f"      🔍 Indexed {min(len(search_entries), 500)} search items")
        
        workbook.close()
        conn.commit()
        
        print(f"✅ Quick processing complete!")
        
        # Show stats
        cursor.execute("SELECT COUNT(*) FROM excel_data")
        data_count = cursor.fetchone()[0]
        
        cursor.execute("SELECT COUNT(*) FROM search_index")
        search_count = cursor.fetchone()[0]
        
        print(f"📊 Total cells: {data_count:,}")
        print(f"🔍 Searchable items: {search_count:,}")
        
    except Exception as e:
        print(f"❌ Error: {e}")
        conn.rollback()
    
    finally:
        conn.close()

def test_search():
    """Test search functionality."""
    print(f"\n🔍 TESTING SEARCH:")
    print("-" * 20)
    
    conn = sqlite3.connect(DATABASE_PATH)
    cursor = conn.cursor()
    
    test_queries = ["ABC", "Construction", "revenue", "profit", "berhad"]
    
    for query in test_queries:
        cursor.execute("""
            SELECT DISTINCT searchable_text 
            FROM search_index 
            WHERE searchable_text LIKE ? 
            LIMIT 3
        """, (f"%{query}%",))
        
        results = cursor.fetchall()
        print(f"'{query}': {len(results)} results")
        for result in results:
            print(f"  - {result[0]}")
    
    conn.close()

if __name__ == "__main__":
    start_time = time.time()
    quick_process()
    test_search()
    print(f"\n⏱️  Total time: {time.time() - start_time:.1f} seconds")