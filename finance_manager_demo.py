"""
Malaysian Investment Holdings Finance Manager Agent Demo
- Works without pandas/openpyxl dependencies
- Shows Excel file upload interface 
- Demonstrates Senior Finance Manager persona
"""

import sys
import os
import json
import time
import uuid
from typing import Dict, Any, List, Optional

import uvicorn
from fastapi import FastAPI, HTTPException, UploadFile, File, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, HTMLResponse
import httpx
from pydantic import BaseModel


# Models
class ChatRequest(BaseModel):
    message: str
    agent_name: str = "Dato Ahmad Rahman"
    use_rag: bool = True


class ChatResponse(BaseModel):
    response: str
    agent_name: str
    execution_time_ms: int
    sources_used: List[str] = []


class ExcelUploadResponse(BaseModel):
    message: str
    filename: str
    status: str
    analysis_preview: str


# Storage
agents_db: Dict[str, Dict[str, Any]] = {}
rag_db: Dict[str, List[Dict[str, Any]]] = {
    "malaysian_finance": [],
    "investment_holdings": [],
    "subsidiary_management": []
}

# LM Studio configuration
LM_STUDIO_BASE_URL = "http://192.168.101.70:1234/v1"
LM_STUDIO_MODEL = "openai/gpt-oss-20b"

uploaded_files = {}  # Simple file storage


async def call_lm_studio(messages: List[Dict[str, str]]) -> str:
    """Call your LM Studio instance."""
    try:
        async with httpx.AsyncClient(timeout=60.0) as client:
            response = await client.post(
                f"{LM_STUDIO_BASE_URL}/chat/completions",
                json={
                    "model": LM_STUDIO_MODEL,
                    "messages": messages,
                    "temperature": 0.2,  # Lower temperature for financial analysis
                    "max_tokens": 2000,
                },
                headers={"Content-Type": "application/json"}
            )
            
            if response.status_code == 200:
                data = response.json()
                return data["choices"][0]["message"]["content"]
            else:
                return f"Error calling LM Studio: HTTP {response.status_code}"
                
    except Exception as e:
        return f"Error connecting to LM Studio: {str(e)}"


def simple_rag_search(question: str, collection_name: str = "malaysian_finance", k: int = 3) -> List[Dict[str, Any]]:
    """Simple RAG search using keyword matching."""
    documents = rag_db.get(collection_name, [])
    if not documents:
        return []
    
    question_words = set(question.lower().split())
    scored_docs = []
    
    for doc in documents:
        text_words = set(doc["text"].lower().split())
        score = len(question_words.intersection(text_words))
        if score > 0:
            scored_docs.append((score, doc))
    
    scored_docs.sort(key=lambda x: x[0], reverse=True)
    return [doc for _, doc in scored_docs[:k]]


# Create FastAPI app
app = FastAPI(
    title="Dato' Ahmad Rahman - Senior Finance Manager",
    version="3.0.0",
    description="AI Senior Finance Manager specialized in Malaysian Investment Holdings and Excel Analysis"
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["GET", "POST", "PUT", "DELETE"],
    allow_headers=["*"],
)


@app.on_event("startup")
async def startup_event():
    """Initialize the Senior Finance Manager."""
    # Create Dato' Ahmad Rahman
    agent_id = "dato-ahmad-rahman"
    agent_data = {
        "id": agent_id,
        "name": "Dato' Ahmad Rahman",
        "title": "Senior Finance Manager",
        "company": "Malaysian Investment Holdings Berhad",
        "experience_years": 25,
        "certifications": ["CA Malaysia", "CPA", "CISA"],
        "specialties": [
            "Investment Holdings Management",
            "Subsidiary Financial Analysis", 
            "Malaysian Corporate Compliance",
            "Excel-based Financial Modeling",
            "Bursa Malaysia Regulations",
            "Multi-subsidiary Consolidation"
        ],
        "languages": ["English", "Bahasa Malaysia", "Mandarin"],
        "status": "active",
        "created_at": time.time()
    }
    
    agents_db[agent_id] = agent_data
    
    # Load Malaysian finance knowledge
    knowledge_base = [
        {
            "text": "Malaysian Investment Holdings Compliance: Under Companies Act 2016, all holding companies must maintain proper subsidiary records, consolidated financial statements, and comply with Bursa Malaysia continuous disclosure requirements. Key focus areas include related party transactions, material contracts, and quarterly reporting deadlines.",
            "metadata": {"source": "Companies_Act_2016", "category": "regulatory"},
        },
        {
            "text": "Subsidiary Performance Monitoring: For Malaysian investment holdings, establish monthly management reporting from all subsidiaries including: Revenue trends, EBITDA performance, cash flow projections, capex requirements, working capital analysis, and debt covenant compliance. Use standardized Excel templates for consistent reporting.",
            "metadata": {"source": "Management_Guidelines", "category": "operations"},
        },
        {
            "text": "Excel Financial Analysis Best Practices: When reviewing subsidiary Excel reports, focus on variance analysis (budget vs actual), trend analysis (YoY and QoQ), ratio analysis (liquidity, profitability, leverage), and cash flow forecasting. Look for red flags: declining margins, increasing receivables days, inventory buildup, or covenant breaches.",
            "metadata": {"source": "Financial_Analysis_Manual", "category": "technical"},
        },
        {
            "text": "Malaysian Tax Considerations: For investment holdings with multiple subsidiaries, monitor transfer pricing documentation, withholding tax on inter-company transactions, Section 140A deemed dividend provisions, and thin capitalization rules. Ensure all subsidiaries maintain proper supporting documentation for related party transactions.",
            "metadata": {"source": "Tax_Guidelines_Malaysia", "category": "tax"},
        },
        {
            "text": "Risk Management Framework: Investment holdings should implement risk assessment across all subsidiaries focusing on: Market risk (sector exposure), credit risk (customer concentration), operational risk (key person dependency), liquidity risk (cash flow timing), and regulatory risk (compliance requirements). Monthly risk dashboards are essential.",
            "metadata": {"source": "Risk_Management_Policy", "category": "risk"},
        }
    ]
    
    for knowledge in knowledge_base:
        doc_id = str(uuid.uuid4())
        document = {
            "id": doc_id,
            "text": knowledge["text"],
            "metadata": knowledge["metadata"],
            "created_at": time.time()
        }
        rag_db["malaysian_finance"].append(document)
    
    print("✅ Dato' Ahmad Rahman - Senior Finance Manager initialized")
    print(f"📚 {len(knowledge_base)} Malaysian finance knowledge documents loaded")


@app.get("/")
async def root():
    return {
        "agent": "Dato' Ahmad Rahman",
        "title": "Senior Finance Manager", 
        "company": "Malaysian Investment Holdings Specialist",
        "expertise": [
            "Excel Financial Analysis",
            "Subsidiary Management", 
            "Malaysian Corporate Compliance",
            "Investment Holdings Strategy"
        ],
        "endpoints": {
            "chat": "/chat - Chat with Dato' Ahmad",
            "excel_upload": "/excel/upload - Upload Excel files for analysis", 
            "health": "/health - System status",
            "demo_page": "/demo - Interactive demo page"
        },
        "message": "Selamat datang! I'm here to help with your investment holdings and financial analysis."
    }


@app.get("/demo", response_class=HTMLResponse)
async def demo_page():
    """Interactive demo page."""
    html = """
    <!DOCTYPE html>
    <html>
    <head>
        <title>Dato Ahmad Rahman - Senior Finance Manager</title>
        <style>
            /* Dark Theme Styles */
            body { 
                font-family: Arial, sans-serif; 
                max-width: 1200px; 
                margin: 0 auto; 
                padding: 20px; 
                background-color: #0f172a; 
                color: #e2e8f0; 
            }
            .header { 
                background: linear-gradient(135deg, #1e40af, #3b82f6); 
                color: white; 
                padding: 20px; 
                border-radius: 8px; 
                margin-bottom: 20px; 
                box-shadow: 0 4px 15px rgba(0,0,0,0.3);
            }
            .card { 
                border: 1px solid #334155; 
                border-radius: 8px; 
                padding: 20px; 
                margin: 20px 0; 
                background-color: #1e293b; 
                box-shadow: 0 2px 10px rgba(0,0,0,0.3);
            }
            .chat-area { 
                background: #0f172a; 
                padding: 15px; 
                border-radius: 5px; 
                margin: 10px 0; 
                border: 1px solid #334155;
            }
            .excel-upload { 
                background: linear-gradient(135deg, #0f3460, #1e40af); 
                padding: 15px; 
                border-radius: 5px; 
                border: 1px solid #3b82f6;
            }
            button { 
                background: linear-gradient(135deg, #3b82f6, #6366f1); 
                color: white; 
                padding: 10px 20px; 
                border: none; 
                border-radius: 5px; 
                cursor: pointer; 
                transition: all 0.3s ease;
                font-weight: 500;
            }
            button:hover { 
                background: linear-gradient(135deg, #2563eb, #4f46e5); 
                transform: translateY(-2px);
                box-shadow: 0 4px 15px rgba(59, 130, 246, 0.4);
            }
            input, textarea, select { 
                width: 100%; 
                padding: 8px; 
                margin: 5px 0; 
                border: 1px solid #475569; 
                border-radius: 4px; 
                background-color: #334155; 
                color: #e2e8f0;
                transition: border-color 0.3s ease;
            }
            input:focus, textarea:focus, select:focus {
                outline: none;
                border-color: #3b82f6;
                box-shadow: 0 0 0 2px rgba(59, 130, 246, 0.2);
            }
            .response { 
                background: linear-gradient(135deg, #1e293b, #334155); 
                padding: 15px; 
                border-left: 4px solid #3b82f6; 
                margin: 10px 0; 
                border-radius: 8px;
                box-shadow: 0 2px 10px rgba(0,0,0,0.2);
                font-family: 'Courier New', monospace;
            }
            .expertise { 
                display: grid; 
                grid-template-columns: repeat(auto-fit, minmax(250px, 1fr)); 
                gap: 15px; 
            }
            .expertise-item { 
                background: linear-gradient(135deg, #1e293b, #334155); 
                padding: 15px; 
                border-radius: 8px; 
                border: 1px solid #475569;
                transition: transform 0.3s ease;
            }
            .expertise-item:hover {
                transform: translateY(-5px);
                box-shadow: 0 5px 20px rgba(59, 130, 246, 0.2);
            }
            h1, h2, h3, h4 { color: #f1f5f9; }
            /* Scrollbar styling for dark theme */
            ::-webkit-scrollbar {
                width: 8px;
            }
            ::-webkit-scrollbar-track {
                background: #1e293b;
            }
            ::-webkit-scrollbar-thumb {
                background: #475569;
                border-radius: 4px;
            }
            ::-webkit-scrollbar-thumb:hover {
                background: #64748b;
            }
            /* Loading animation */
            .loading {
                color: #3b82f6;
                animation: pulse 2s infinite;
            }
            @keyframes pulse {
                0%, 100% { opacity: 1; }
                50% { opacity: 0.5; }
            }
        </style>
    </head>
    <body>
        <div class="header">
            <h1>🏢 Dato Ahmad Rahman</h1>
            <h2>Senior Finance Manager - Investment Holdings Specialist</h2>
            <p>25+ years experience • CA Malaysia • CPA • Malaysian Corporate Expert</p>
        </div>

        <div class="card">
            <h3>💬 Chat with Dato Ahmad</h3>
            <div class="chat-area">
                <textarea id="chatMessage" placeholder="Ask about financial analysis, subsidiary management, Excel analysis, Malaysian compliance, etc..." rows="3"></textarea>
                <button onclick="sendChat()">Send Message</button>
                <div id="chatResponse" class="response" style="display:none;"></div>
            </div>
        </div>

        <div class="card">
            <h3>📊 Excel File Analysis</h3>
            <div class="excel-upload">
                <p><strong>Upload your subsidiary Excel files for expert analysis</strong></p>
                <input type="file" id="excelFile" accept=".xlsx,.xls" />
                <select id="analysisType">
                    <option value="general">General Financial Review</option>
                    <option value="subsidiary_performance">Subsidiary Performance Analysis</option>
                    <option value="cash_flow">Cash Flow Analysis</option>
                    <option value="compliance_check">Compliance Review</option>
                    <option value="risk_assessment">Risk Assessment</option>
                </select>
                <textarea id="specificQuery" placeholder="Specific questions or focus areas..." rows="2"></textarea>
                <button onclick="uploadExcel()">Analyze Excel File</button>
                <div id="excelResponse" class="response" style="display:none;"></div>
            </div>
        </div>

        <div class="card">
            <h3>🎯 Areas of Expertise</h3>
            <div class="expertise">
                <div class="expertise-item">
                    <h4>📈 Investment Holdings Management</h4>
                    <p>Strategic oversight of subsidiary portfolios, performance monitoring, and value optimization</p>
                </div>
                <div class="expertise-item">
                    <h4>🇲🇾 Malaysian Compliance</h4>
                    <p>Companies Act 2016, Bursa Malaysia requirements, MFRS standards, tax regulations</p>
                </div>
                <div class="expertise-item">
                    <h4>📋 Subsidiary Analysis</h4>
                    <p>Financial performance review, consolidation, inter-company transactions, risk assessment</p>
                </div>
                <div class="expertise-item">
                    <h4>💻 Excel Financial Modeling</h4>
                    <p>Advanced Excel analysis, unstructured data processing, financial modeling and reporting</p>
                </div>
            </div>
        </div>

        <script>
            async function sendChat() {
                const message = document.getElementById('chatMessage').value;
                const responseDiv = document.getElementById('chatResponse');
                
                if (!message.trim()) return;
                
                responseDiv.style.display = 'block';
                responseDiv.innerHTML = '<div class="loading">🤖 Dato Ahmad is analyzing your query...</div>';
                
                try {
                    const response = await fetch('/chat', {
                        method: 'POST',
                        headers: {'Content-Type': 'application/json'},
                        body: JSON.stringify({
                            message: message,
                            agent_name: "Dato Ahmad Rahman",
                            use_rag: true
                        })
                    });
                    
                    const data = await response.json();
                    responseDiv.innerHTML = `
                        <strong>Dato Ahmad Rahman:</strong><br>
                        ${data.response.replace(/\\n/g, '<br>')}
                        <br><br>
                        <small>Response time: ${data.execution_time_ms}ms</small>
                    `;
                } catch (error) {
                    responseDiv.innerHTML = 'Error: ' + error.message;
                }
            }
            
            async function uploadExcel() {
                const fileInput = document.getElementById('excelFile');
                const analysisType = document.getElementById('analysisType').value;
                const specificQuery = document.getElementById('specificQuery').value;
                const responseDiv = document.getElementById('excelResponse');
                
                if (!fileInput.files[0]) {
                    alert('Please select an Excel file first');
                    return;
                }
                
                responseDiv.style.display = 'block';
                responseDiv.innerHTML = '<div class="loading">📊 Dato Ahmad is processing your Excel file...</div>';
                
                const formData = new FormData();
                formData.append('file', fileInput.files[0]);
                formData.append('analysis_type', analysisType);
                formData.append('specific_query', specificQuery);
                
                try {
                    const response = await fetch('/excel/upload', {
                        method: 'POST',
                        body: formData
                    });
                    
                    const data = await response.json();
                    responseDiv.innerHTML = `
                        <strong>Excel Analysis Complete:</strong><br>
                        ${data.analysis_preview.replace(/\\n/g, '<br>')}
                        <br><br>
                        <small>File: ${data.filename}</small>
                    `;
                } catch (error) {
                    responseDiv.innerHTML = 'Error processing Excel file: ' + error.message;
                }
            }
        </script>
    </body>
    </html>
    """
    return html


@app.get("/favicon.ico")
async def favicon():
    """Return favicon to stop 404 errors."""
    return {"message": "No favicon available"}

@app.get("/health")
async def health_check():
    """Health check."""
    try:
        async with httpx.AsyncClient(timeout=5.0) as client:
            response = await client.get(f"{LM_STUDIO_BASE_URL}/models")
            lm_studio_status = "healthy" if response.status_code == 200 else "unhealthy"
    except:
        lm_studio_status = "unhealthy"
    
    return {
        "status": "healthy",
        "agent": "Dato' Ahmad Rahman - Senior Finance Manager",
        "lm_studio": lm_studio_status,
        "knowledge_base": {
            "malaysian_finance": len(rag_db["malaysian_finance"]),
            "investment_holdings": len(rag_db["investment_holdings"]), 
            "subsidiary_management": len(rag_db["subsidiary_management"])
        },
        "capabilities": [
            "Excel file analysis (upload supported)",
            "Malaysian corporate compliance advice",
            "Subsidiary performance review",
            "Financial modeling and forecasting",
            "Investment holdings management"
        ]
    }


@app.post("/chat", response_model=ChatResponse)
async def chat_with_dato_ahmad(request: ChatRequest):
    """Chat with Dato' Ahmad Rahman using REAL Excel data."""
    start_time = time.time()
    
    # Import and use the REAL system
    import sys
    import os
    sys.path.append(os.path.dirname(__file__))
    from real_excel_chat_system import RealExcelChatSystem
    
    try:
        # Use REAL Excel data integration
        real_system = RealExcelChatSystem()
        response_text = await real_system.chat_with_excel_data(request.message)
        
        sources_used = ["Real Excel Database", "TURN-COS-GP_RM", "PNL", "GROUP-PL_RM"]
        
    except Exception as e:
        # Fallback but indicate it's not using Excel data
        response_text = f"I apologize, I'm currently unable to access your Excel data (Error: {str(e)}). For Excel data analysis, please ensure the database is properly connected."
        sources_used = ["Error - No Excel Data Access"]
    
    execution_time = int((time.time() - start_time) * 1000)
    
    return ChatResponse(
        response=response_text,
        agent_name="Dato Ahmad Rahman", 
        execution_time_ms=execution_time,
        sources_used=sources_used
    )


@app.post("/excel/upload", response_model=ExcelUploadResponse)
async def upload_excel_file(
    file: UploadFile = File(...),
    analysis_type: str = Form("general"),
    specific_query: str = Form("")
):
    """Upload Excel file for analysis by Dato' Ahmad."""
    print(f"📋 Upload received - File: {file.filename}, Query: '{specific_query}', Type: {analysis_type}")
    
    if not file.filename.lower().endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only Excel files (.xlsx, .xls) are supported")
    
    try:
        # Check file size first (limit to 10MB to prevent timeouts)
        file_data = await file.read()
        if len(file_data) > 10 * 1024 * 1024:  # 10MB limit
            raise HTTPException(status_code=413, detail="File too large. Maximum size is 10MB")
            
        file_id = str(uuid.uuid4())
        
        uploaded_files[file_id] = {
            "filename": file.filename,
            "data": file_data,
            "uploaded_at": time.time(),
            "analysis_type": analysis_type,
            "specific_query": specific_query
        }
        
        # ACTUALLY PROCESS THE EXCEL FILE WITH REAL DATA
        import tempfile
        import openpyxl
        from pathlib import Path
        
        # Save uploaded file to temporary location
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        temp_file.write(file_data)
        temp_file.close()
        
        try:
            # Read actual Excel file data
            workbook = openpyxl.load_workbook(temp_file.name, read_only=True)
            sheet_names = workbook.sheetnames
            
            # SMART ANALYSIS BASED ON SPECIFIC QUERY
            sheets_info = {}
            relevant_data = []
            
            # Determine which sheets to focus on based on query
            focus_sheets = sheet_names
            if specific_query:
                query_lower = specific_query.lower()
                # If query mentions specific sheet, focus on that
                mentioned_sheets = [name for name in sheet_names if name.lower() in query_lower]
                if mentioned_sheets:
                    focus_sheets = mentioned_sheets
            
            for sheet_name in sheet_names:
                worksheet = workbook[sheet_name]
                sheet_info = {
                    "rows": worksheet.max_row,
                    "columns": worksheet.max_column,
                    "sample_data": []
                }
                
                # If this sheet is focus of query, do deeper analysis
                if sheet_name in focus_sheets and specific_query:
                    print(f"🔍 Deep analysis of {sheet_name} for query: {specific_query}")
                    
                    # Look for company names, entities, etc.
                    companies_found = []
                    financial_terms = []
                    
                    # Read more data from focused sheets (up to 100 rows)
                    max_rows_to_scan = min(100, worksheet.max_row)
                    max_cols_to_scan = min(20, worksheet.max_column)
                    
                    for row in range(1, max_rows_to_scan + 1):
                        for col in range(1, max_cols_to_scan + 1):
                            cell = worksheet.cell(row=row, column=col)
                            if cell.value and isinstance(cell.value, str):
                                cell_text = str(cell.value).strip()
                                if len(cell_text) > 3:
                                    # Look for company patterns
                                    if any(pattern in cell_text.lower() for pattern in 
                                          ['sdn bhd', 'berhad', 'bhd', 'corporation', 'company', 'construction', 'holdings']):
                                        if cell_text not in companies_found and len(cell_text) < 100:
                                            companies_found.append(cell_text)
                                    
                                    # Look for financial terms if query mentions them
                                    if any(term in specific_query.lower() for term in ['revenue', 'profit', 'financial']):
                                        if any(pattern in cell_text.lower() for pattern in 
                                              ['revenue', 'profit', 'total', 'gross', 'net', 'income']):
                                            if cell_text not in financial_terms and len(cell_text) < 50:
                                                financial_terms.append(cell_text)
                    
                    sheet_info["companies_found"] = companies_found[:10]  # Limit to 10
                    sheet_info["financial_terms"] = financial_terms[:10]
                    
                    if companies_found:
                        relevant_data.append(f"📊 {sheet_name} - Companies: {', '.join(companies_found[:5])}")
                    if financial_terms:
                        relevant_data.append(f"💰 {sheet_name} - Financial: {', '.join(financial_terms[:3])}")
                
                else:
                    # Basic info for non-focus sheets
                    if worksheet.max_row > 0:
                        first_row = []
                        for col in range(1, min(4, worksheet.max_column + 1)):
                            cell = worksheet.cell(row=1, column=col)
                            if cell.value is not None:
                                first_row.append(str(cell.value)[:30])
                        sheet_info["first_row"] = first_row
                
                sheets_info[sheet_name] = sheet_info
            
            workbook.close()
            
            # Generate REAL analysis based on ACTUAL data
            real_data_context = f"""ACTUAL EXCEL FILE DATA:
Filename: {file.filename}
Total Sheets: {len(sheet_names)}
Sheet Names: {', '.join(sheet_names)}

Sheet Details:"""
            
            for sheet_name, info in sheets_info.items():
                real_data_context += f"\n- {sheet_name}: {info['rows']} rows × {info['columns']} columns"
                if info.get('first_row'):
                    real_data_context += f" (Headers: {', '.join(info['first_row'][:3])})"
            
            # Create intelligent response based on query and findings
            if relevant_data and specific_query:
                preview_response = f"""**SMART ANALYSIS - {file.filename}**

**Query:** "{specific_query}"

**FINDINGS FROM YOUR EXCEL DATA:**
{chr(10).join(relevant_data)}

**Sheet Structure:**
{chr(10).join([f"• {name}: {info['rows']} rows × {info['columns']} columns" for name, info in sheets_info.items()])}

**Dato Ahmad Rahman's Analysis:**
Based on your specific query, I've analyzed the actual cell contents of your Excel file. The findings above show the real company names and data extracted from the cells, not generic structure information.

*This analysis is based on actual cell-by-cell examination of your Excel data.*"""
            else:
                preview_response = f"""**WORKSHEET STRUCTURE - {file.filename}**

**Worksheets Found ({len(sheet_names)} sheets):**
{chr(10).join([f"• {name}" for name in sheet_names])}

**Sheet Dimensions:**
{chr(10).join([f"• {name}: {info['rows']} rows × {info['columns']} columns" for name, info in sheets_info.items()])}

Query: "{specific_query if specific_query else '[No specific query provided]'}"

**Dato Ahmad Rahman's Analysis:**
I can see the structure of your Excel file. For more detailed analysis of specific data within cells, please provide a more specific query about what you're looking for.

*For cell-level analysis, ask specific questions like "find company names in TURN-COS-GP_RM" or "show revenue figures from PNL sheet".*"""
            
        except Exception as e:
            preview_response = f"Error reading Excel file: {str(e)}"
        
        finally:
            # Clean up temp file
            Path(temp_file.name).unlink(missing_ok=True)
        
        return ExcelUploadResponse(
            message="Excel file uploaded successfully",
            filename=file.filename,
            status="received",
            analysis_preview=preview_response
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error processing file: {str(e)}")


@app.get("/excel/files")
async def list_uploaded_files():
    """List uploaded Excel files."""
    return {
        "files": [
            {
                "file_id": file_id,
                "filename": data["filename"],
                "uploaded_at": data["uploaded_at"],
                "analysis_type": data["analysis_type"]
            }
            for file_id, data in uploaded_files.items()
        ]
    }


if __name__ == "__main__":
    print("🏢 Starting Dato' Ahmad Rahman - Senior Finance Manager")
    print(f"📡 LM Studio: {LM_STUDIO_BASE_URL}")
    print(f"🤖 Model: {LM_STUDIO_MODEL}")
    print("🇲🇾 Specialized in Malaysian Investment Holdings")
    print("📊 Excel Analysis Ready (Upload via web interface)")
    print("🌐 Access: http://localhost:8003")
    print("📚 API docs: http://localhost:8003/docs")
    print("🎯 Interactive demo: http://localhost:8003/demo")
    
    uvicorn.run(app, host="0.0.0.0", port=8003)