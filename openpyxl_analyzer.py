"""
Real Excel File Analyzer using openpyxl (the best Excel library!)
Dato' Ahmad Rahman will analyze your actual multi_sheet.xlsx file.
"""

import sys
import os
import json
import asyncio
from pathlib import Path

# Import openpyxl - the BEST Excel processing library
import openpyxl
from openpyxl import load_workbook

import httpx

# Your file path
EXCEL_FILE_PATH = "/mnt/c/Users/nvntr/Documents/ai_agent_platform/data/input/multi_sheet.xlsx"
LM_STUDIO_BASE_URL = "http://192.168.101.70:1234/v1"
LM_STUDIO_MODEL = "openai/gpt-oss-20b"


async def call_lm_studio(messages):
    """Call your LM Studio."""
    try:
        async with httpx.AsyncClient(timeout=120.0) as client:
            response = await client.post(
                f"{LM_STUDIO_BASE_URL}/chat/completions",
                json={
                    "model": LM_STUDIO_MODEL,
                    "messages": messages,
                    "temperature": 0.1,  # Very low for precise analysis
                    "max_tokens": 3000,
                },
                headers={"Content-Type": "application/json"}
            )
            
            if response.status_code == 200:
                data = response.json()
                return data["choices"][0]["message"]["content"]
            else:
                return f"Error calling LM Studio: HTTP {response.status_code}"
                
    except Exception as e:
        return f"Error: {str(e)}"


def analyze_real_excel_with_openpyxl():
    """Use openpyxl to read your ACTUAL Excel file - the best way!"""
    if not os.path.exists(EXCEL_FILE_PATH):
        return {"error": f"File not found: {EXCEL_FILE_PATH}"}
    
    try:
        print(f"📖 Reading your REAL Excel file with openpyxl: {EXCEL_FILE_PATH}")
        
        # Load the workbook (read-only for performance)
        workbook = load_workbook(EXCEL_FILE_PATH, read_only=True)
        
        real_data = {
            "filename": "multi_sheet.xlsx",
            "total_sheets": len(workbook.sheetnames),
            "sheet_names": workbook.sheetnames,
            "sheets_analysis": {}
        }
        
        print(f"📋 Found {len(workbook.sheetnames)} sheets: {workbook.sheetnames}")
        
        # Analyze each sheet with REAL data
        for sheet_name in workbook.sheetnames:
            print(f"   🔍 Analyzing sheet: '{sheet_name}'")
            
            try:
                worksheet = workbook[sheet_name]
                
                # Get actual dimensions
                max_row = worksheet.max_row
                max_col = worksheet.max_column
                
                # Extract REAL data (first 20 rows for detailed analysis)
                real_cell_data = []
                numeric_values = []
                text_values = []
                date_values = []
                
                for row_num in range(1, min(21, max_row + 1)):  # First 20 rows
                    row_data = {}
                    for col_num in range(1, min(16, max_col + 1)):  # First 15 columns
                        cell = worksheet.cell(row=row_num, column=col_num)
                        cell_value = cell.value
                        col_letter = openpyxl.utils.get_column_letter(col_num)
                        
                        if cell_value is not None:
                            if isinstance(cell_value, (int, float)):
                                numeric_values.append(float(cell_value))
                                row_data[f"{col_letter}{row_num}"] = cell_value
                            elif isinstance(cell_value, str) and cell_value.strip():
                                text_values.append(cell_value.strip()[:100])  # Limit length
                                row_data[f"{col_letter}{row_num}"] = cell_value.strip()[:50]
                            else:
                                row_data[f"{col_letter}{row_num}"] = str(cell_value)[:50]
                        else:
                            row_data[f"{col_letter}{row_num}"] = ""
                    
                    if any(row_data.values()):  # Only add non-empty rows
                        real_cell_data.append(row_data)
                
                # Analyze the REAL content
                financial_terms = []
                malaysian_entities = []
                currency_amounts = []
                dates_found = []
                
                for text in text_values:
                    text_lower = text.lower()
                    
                    # Financial terms
                    if any(term in text_lower for term in ['revenue', 'profit', 'loss', 'assets', 'cash', 'expense', 'income', 'balance', 'total']):
                        if text not in financial_terms:
                            financial_terms.append(text)
                    
                    # Malaysian business entities  
                    if any(term in text_lower for term in ['sdn bhd', 'berhad', 'bhd', 'malaysia', 'kuala lumpur']):
                        if text not in malaysian_entities:
                            malaysian_entities.append(text)
                    
                    # Currency mentions
                    if any(term in text_lower for term in ['rm ', 'ringgit', 'usd', 'million', 'thousand']):
                        if text not in currency_amounts:
                            currency_amounts.append(text)
                
                # Statistical analysis of numeric data
                numeric_stats = {}
                if numeric_values:
                    numeric_stats = {
                        "count": len(numeric_values),
                        "sum": sum(numeric_values),
                        "average": sum(numeric_values) / len(numeric_values),
                        "max": max(numeric_values),
                        "min": min(numeric_values),
                        "large_values": [v for v in numeric_values if v > 1000000],  # Values over 1M
                        "negative_values": [v for v in numeric_values if v < 0],
                    }
                
                real_data["sheets_analysis"][sheet_name] = {
                    "dimensions": {"rows": max_row, "columns": max_col},
                    "real_cell_data": real_cell_data[:10],  # First 10 rows of actual data
                    "numeric_statistics": numeric_stats,
                    "financial_terms_found": financial_terms[:10],
                    "malaysian_entities_found": malaysian_entities[:5],
                    "currency_mentions": currency_amounts[:5],
                    "total_non_empty_cells": len([v for row in real_cell_data for v in row.values() if v]),
                    "content_types": {
                        "numeric_cells": len(numeric_values),
                        "text_cells": len(text_values),
                        "empty_cells": (min(20, max_row) * min(15, max_col)) - len(numeric_values) - len(text_values)
                    }
                }
                
            except Exception as e:
                real_data["sheets_analysis"][sheet_name] = {"error": f"Error analyzing sheet: {str(e)}"}
        
        workbook.close()
        return real_data
        
    except Exception as e:
        return {"error": f"Failed to read Excel file: {str(e)}"}


async def get_dato_ahmad_real_analysis(excel_data):
    """Get Dato' Ahmad's analysis of your ACTUAL Excel data."""
    
    # Build detailed context with REAL data from your file
    context = f"""ACTUAL EXCEL FILE DATA ANALYSIS - REAL DATA FROM USER'S MULTI_SHEET.XLSX

File: {excel_data['filename']}
Total Sheets: {excel_data['total_sheets']}
Sheet Names: {', '.join(excel_data['sheet_names'])}

DETAILED ANALYSIS OF REAL DATA:
"""
    
    # Add analysis of each sheet with actual data
    for sheet_name, analysis in excel_data['sheets_analysis'].items():
        if 'error' not in analysis:
            context += f"\n=== SHEET: {sheet_name} (ACTUAL DATA) ===\n"
            context += f"Real Dimensions: {analysis['dimensions']['rows']} rows × {analysis['dimensions']['columns']} columns\n"
            
            # Add actual cell data
            if analysis['real_cell_data']:
                context += "REAL CELL VALUES FROM YOUR FILE:\n"
                for i, row_data in enumerate(analysis['real_cell_data'][:5], 1):
                    non_empty = {k: v for k, v in row_data.items() if v and str(v).strip()}
                    if non_empty:
                        context += f"Row {i}: {str(non_empty)[:200]}...\n"
            
            # Add numeric analysis
            if analysis['numeric_statistics']['count'] > 0:
                stats = analysis['numeric_statistics']
                context += f"NUMERIC ANALYSIS:\n"
                context += f"- Total numbers: {stats['count']}\n"
                context += f"- Sum: {stats['sum']:,.2f}\n"
                context += f"- Average: {stats['average']:,.2f}\n"
                context += f"- Range: {stats['min']:,.2f} to {stats['max']:,.2f}\n"
                if stats['large_values']:
                    context += f"- Large values (>1M): {[f'{v:,.0f}' for v in stats['large_values'][:5]]}\n"
            
            # Add content findings
            if analysis['financial_terms_found']:
                context += f"FINANCIAL TERMS IN YOUR DATA: {', '.join(analysis['financial_terms_found'][:5])}\n"
            
            if analysis['malaysian_entities_found']:
                context += f"MALAYSIAN ENTITIES IN YOUR DATA: {', '.join(analysis['malaysian_entities_found'])}\n"
            
            if analysis['currency_mentions']:
                context += f"CURRENCY REFERENCES: {', '.join(analysis['currency_mentions'])}\n"
            
            context += f"Content Distribution: {analysis['content_types']['numeric_cells']} numbers, {analysis['content_types']['text_cells']} text entries\n\n"
    
    # Dato' Ahmad's comprehensive analysis prompt
    prompt = f"""As Dato' Ahmad Rahman, I am now analyzing the USER'S ACTUAL Excel file data. This is REAL data extracted directly from their multi_sheet.xlsx file using openpyxl.

{context}

Based on this REAL data from the user's actual Excel file, I need to provide:

1. **ACTUAL DATA ASSESSMENT**: What specific financial information do I see in their real data?

2. **BUSINESS CONTEXT ANALYSIS**: Based on the entity names and financial terms found, what type of business/projects are these?

3. **SHEET-BY-SHEET RECOMMENDATIONS**: Which sheets contain the most valuable data for financial reporting?

4. **SPECIFIC DATA EXTRACTION PLAN**: Exactly how to extract key metrics from these actual sheets

5. **MALAYSIAN COMPLIANCE INSIGHTS**: Compliance considerations based on the actual entity types found

6. **ACTIONABLE NEXT STEPS**: Concrete steps to get the information needed for their board report

Remember: I'm analyzing REAL data from their actual Excel file. I should reference specific values, sheet names, and entities I can see in their data."""
    
    messages = [
        {
            "role": "system",
            "content": """You are Dato' Ahmad Rahman analyzing REAL Excel data from a user's actual investment holding company file. 

Provide specific, actionable analysis based on the actual data shown. Reference specific values, sheet names, and business entities you can see in the real data provided. Be practical and detailed."""
        },
        {
            "role": "user",
            "content": prompt
        }
    ]
    
    print("\n💭 Dato' Ahmad is analyzing your REAL Excel data...")
    analysis = await call_lm_studio(messages)
    return analysis


async def main():
    print("🏢 DATO' AHMAD RAHMAN - REAL EXCEL ANALYSIS WITH OPENPYXL")
    print("=" * 65)
    print(f"📁 Analyzing your actual file: {EXCEL_FILE_PATH}")
    print("🔧 Using openpyxl - the best Excel processing library!")
    print()
    
    # Analyze real Excel data
    print("📊 READING ACTUAL DATA FROM YOUR EXCEL FILE...")
    excel_data = analyze_real_excel_with_openpyxl()
    
    if 'error' in excel_data:
        print(f"❌ Error: {excel_data['error']}")
        return
    
    print(f"✅ Successfully analyzed {excel_data['total_sheets']} sheets with REAL data!")
    
    # Show quick summary
    for sheet_name, analysis in excel_data['sheets_analysis'].items():
        if 'error' not in analysis:
            print(f"   📋 {sheet_name}: {analysis['dimensions']['rows']} rows, {analysis['content_types']['numeric_cells']} numbers")
    
    # Get Dato' Ahmad's analysis of real data  
    print("\n👔 DATO' AHMAD'S ANALYSIS OF YOUR ACTUAL DATA:")
    print("-" * 55)
    
    analysis = await get_dato_ahmad_real_analysis(excel_data)
    print(analysis)
    
    print("\n" + "=" * 65)
    print("📋 REAL DATA ANALYSIS COMPLETE")
    print("✅ This analysis is based on your actual Excel file content!")
    print("💬 Dato' Ahmad now has access to your real business data")


if __name__ == "__main__":
    asyncio.run(main())